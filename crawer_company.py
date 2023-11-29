import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
os.chdir('C:\sunthing\python\crawer')

import openpyxl
wb=openpyxl.load_workbook('total.xlsx')

names=wb.sheetnames
s1=wb.active
s2=wb['歌曲']
print(names)
print(s1, s1.max_row, s1.max_column)
print(s2, s2.max_row, s2.max_column)

#同等於print(s2['B2'].value)
print(s2.cell(2,2).value)

#問題:iter_rows跟iter_cols的差別在哪
v=s2.iter_cols(min_row=2, min_col=4, max_row=121, max_col=4)
print(v)

user={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'}
data_list=[]
for i in v:
    for j in i:
        data={}
        url='https://www.google.com/search?q='
        key=j.value
        print(j.value)
        data['歌曲名稱']=j.value
        url=url+key+'&sca_esv=586268016&tbm=vid&sxsrf=AM9HkKmEtbImaJJTxTPg7IamFCjAD8TwUA:1701259007995&source=lnms&sa=X&ved=2ahUKEwiwod7Wk-mCAxXKffUHHelgCIMQ_AUoAXoECAEQAw&biw=899&bih=747&dpr=1.25'
        responses=requests.get(url, headers=user)
        soup=BeautifulSoup(responses.text, 'html.parser')
        webanswer=soup.find('div',class_="c8rnLc flgn0c zCaigb")
        if(webanswer):
            webanswer=webanswer.select_one('span')
            print(webanswer.text)
            data['歌曲長度']=webanswer.text

        data_list.append(data)

df=pd.DataFrame(data_list)
df.to_excel('time.xlsx', index= False, engine='openpyxl')






